import csv
import copy
import os
from datetime import datetime, date
from decimal import Decimal

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QTextEdit, QTabWidget, QLabel, QPushButton, QFileDialog, QMessageBox,
    QComboBox, QApplication, QMenu
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont, QColor

from infrastructure.i18n import I18N


_CHANGED_COLOR = QColor("#fff3cd")
_NORMAL_BG = QColor("#ffffff")
_ALT_BG = QColor("#f5f5f5")
_NULL_COLOR = QColor("#888")


class _EditableTable(QTableWidget):
    paste_occurred = Signal(int, int, list)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_V and event.modifiers() & Qt.ControlModifier:
            self._handle_paste()
        else:
            super().keyPressEvent(event)

    def _handle_paste(self):
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if not text:
            return
        rows_text = text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        data = [r.split('\t') for r in rows_text]
        while data and data[-1] == ['']:
            data.pop()
        if not data or not data[0]:
            return
        current = self.currentIndex()
        start_row = current.row() if current.isValid() else 0
        start_col = current.column() if current.isValid() else 0
        self.paste_occurred.emit(start_row, start_col, data)


def _quote_sql(val):
    if val is None:
        return "NULL"
    if isinstance(val, (int, float)):
        return str(val)
    return "'" + str(val).replace("'", "''") + "'"


_NUMERIC_TYPES_SQL = {
    "INT", "BIGINT", "SMALLINT", "TINYINT", "BIT",
    "DECIMAL", "NUMERIC", "FLOAT", "REAL",
    "MONEY", "SMALLMONEY"
}


def _is_numeric_value(val) -> bool:
    return isinstance(val, (int, float, Decimal)) and not isinstance(val, bool)


def _format_number(val) -> str:
    return str(val)


def _format_date(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")


def _format_display_value(val):
    if val is None:
        return "NULL"
    if isinstance(val, bool):
        return "1" if val else "0"
    if _is_numeric_value(val):
        return _format_number(val)
    if isinstance(val, (datetime, date)):
        return _format_date(val)
    return str(val)


class ResultPanel(QWidget):
    status_message = Signal(str, int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._last_columns: list[str] = []
        self._last_rows: list[list] = []
        self._original_rows: list[list] = []
        self._changed_cells: set[tuple[int, int]] = set()
        self._page_size = 100
        self._current_page = 0
        self._total_pages = 0
        self._editable = False
        self._table_name = ""
        self._original_row_count = 0
        self._column_types: dict[str, str] = {}
        self._adapter = None
        self._italic_font = QFont("Consolas", 9)
        self._italic_font.setItalic(True)
        self._normal_font = QFont("Consolas", 9)
        self._build_ui()

    def set_adapter(self, adapter):
        self._adapter = adapter

    def _fetch_column_types(self):
        self._column_types = {}
        if not self._adapter or not self._adapter.is_connected() or not self._table_name:
            return

        name = self._table_name.strip("[]")
        schema = None
        table = name
        if "." in name:
            parts = name.split(".", 1)
            schema = parts[0].strip("[]")
            table = parts[1].strip("[]")

        try:
            cols = self._adapter.get_table_columns(table, schema)
            self._column_types = {c.name: c.data_type for c in cols}
        except Exception:
            self._column_types = {}

    @staticmethod
    def _parse_br_number(s: str):
        s = s.strip()
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(".", "")
        try:
            return int(s)
        except ValueError:
            return float(s)

    def _convert_save_value(self, col_name: str, raw) -> object:
        raw_type = self._column_types.get(col_name, "").upper()
        if raw is None or str(raw).strip().upper() == "NULL":
            return None
        if raw_type in ("DATE", "DATETIME", "DATETIME2", "SMALLDATETIME"):
            from datetime import datetime, date
            val = str(raw).strip().strip("'\"")
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f",
                        "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S",
                        "%Y%m%d", "%d-%m-%Y"):
                try:
                    dt = datetime.strptime(val, fmt)
                    if raw_type == "DATE":
                        return dt.date()
                    return dt
                except ValueError:
                    continue
            return raw
        if raw_type in ("INT", "BIGINT", "SMALLINT", "TINYINT"):
            try:
                return int(str(raw).strip())
            except ValueError:
                try:
                    return int(self._parse_br_number(str(raw)))
                except (ValueError, TypeError):
                    return raw
        if raw_type in ("DECIMAL", "NUMERIC", "FLOAT", "REAL", "MONEY"):
            try:
                return float(str(raw).strip())
            except ValueError:
                try:
                    return self._parse_br_number(str(raw))
                except (ValueError, TypeError):
                    return raw
        return raw

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        self.tabs = QTabWidget()

        self.save_btn = QPushButton(I18N.result_panel["save_changes"])
        self.save_btn.setVisible(False)
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #ffc107; color: #333; padding: 8px 14px;
                font-size: 10px; border-radius: 4px; font-weight: bold;
            }
            QPushButton:hover { background-color: #e0a800; }
            QPushButton:disabled { color: #aaa; background-color: #eee; }
        """)
        self.save_btn.clicked.connect(self._on_save_changes)

        self.export_btn = QPushButton(I18N.result_panel["export_csv"])
        self.export_btn.setEnabled(False)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8; color: white; padding: 8px 12px;
                font-size: 10px; border-radius: 4px; font-weight: bold;
            }
            QPushButton:hover { background-color: #138496; }
            QPushButton:disabled { color: #aaa; background-color: #ccc; }
        """)
        self.export_menu = QMenu(self)
        self.export_menu.addAction("CSV (.csv)", self._export_csv)
        self.export_menu.addAction("Excel (.xlsx)", self._export_excel)
        self.export_btn.setMenu(self.export_menu)

        layout.addWidget(self.tabs)

        self.table = _EditableTable()
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #fff;
                gridline-color: #ddd;
                font-size: 10px;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 4px;
                font-weight: bold;
                border: 1px solid #ddd;
            }
        """)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.cellChanged.connect(self._on_cell_changed)
        self.table.paste_occurred.connect(self._on_paste)

        self.message_display = QTextEdit()
        self.message_display.setReadOnly(True)
        self.message_display.setFont(QFont("Consolas", 10))
        self.message_display.setStyleSheet("""
            QTextEdit {
                background-color: #1e1e1e;
                color: #d4d4d4;
                border: none;
                padding: 8px;
            }
        """)

        self.tabs.addTab(self.table, I18N.result_panel["tab_results"])
        self.tabs.addTab(self.message_display, I18N.result_panel["tab_messages"])

        pagination_layout = QHBoxLayout()
        pagination_layout.setContentsMargins(0, 0, 0, 0)

        self.page_first_btn = QPushButton(I18N.result_panel["page_first"])
        self.page_first_btn.setStyleSheet("padding: 2px 8px; font-size: 10px;")
        self.page_first_btn.setEnabled(False)
        self.page_first_btn.clicked.connect(lambda: self._go_to_page(0))

        self.page_prev_btn = QPushButton(I18N.result_panel["page_prev"])
        self.page_prev_btn.setStyleSheet("padding: 2px 8px; font-size: 10px;")
        self.page_prev_btn.setEnabled(False)
        self.page_prev_btn.clicked.connect(self._prev_page)

        self.page_label = QLabel("")
        self.page_label.setStyleSheet("font-size: 10px; padding: 0 8px;")

        self.page_next_btn = QPushButton(I18N.result_panel["page_next"])
        self.page_next_btn.setStyleSheet("padding: 2px 8px; font-size: 10px;")
        self.page_next_btn.setEnabled(False)
        self.page_next_btn.clicked.connect(self._next_page)

        self.page_last_btn = QPushButton(I18N.result_panel["page_last"])
        self.page_last_btn.setStyleSheet("padding: 2px 8px; font-size: 10px;")
        self.page_last_btn.setEnabled(False)
        self.page_last_btn.clicked.connect(lambda: self._go_to_page(self._total_pages - 1))

        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["50", "100", "200", "500", "1000"])
        self.page_size_combo.setCurrentText(str(self._page_size))
        self.page_size_combo.setStyleSheet("font-size: 10px; padding: 1px 4px;")
        self.page_size_combo.currentTextChanged.connect(self._on_page_size_changed)

        pagination_layout.addStretch()
        pagination_layout.addWidget(self.page_first_btn)
        pagination_layout.addWidget(self.page_prev_btn)
        pagination_layout.addWidget(self.page_label)
        pagination_layout.addWidget(self.page_next_btn)
        pagination_layout.addWidget(self.page_last_btn)
        pagination_layout.addSpacing(16)
        pagination_layout.addWidget(QLabel(I18N.result_panel["page_size_label"]))
        pagination_layout.addWidget(self.page_size_combo)
        pagination_layout.addStretch()

        layout.addLayout(pagination_layout)

    def _prev_page(self):
        if self._current_page > 0:
            self._go_to_page(self._current_page - 1)

    def _next_page(self):
        if self._current_page < self._total_pages - 1:
            self._go_to_page(self._current_page + 1)

    def _go_to_page(self, page: int):
        self._current_page = page
        self._render_page()

    def _on_page_size_changed(self, value: str):
        try:
            new_size = int(value)
            if new_size < 1:
                return
            self._page_size = new_size
            self._update_pagination()
            self._go_to_page(0)
        except ValueError:
            pass

    def _update_pagination(self):
        total = len(self._last_rows)
        self._total_pages = max(1, (total + self._page_size - 1) // self._page_size)
        if self._current_page >= self._total_pages:
            self._current_page = self._total_pages - 1

    def _global_row(self, local_row: int) -> int:
        return self._current_page * self._page_size + local_row

    def _render_page(self):
        total = len(self._last_rows)
        if total == 0:
            self.table.setColumnCount(len(self._last_columns))
            self.table.setHorizontalHeaderLabels(self._last_columns)
            self.table.setRowCount(0)
            self._resize_columns_smart()
            self._update_nav_buttons()
            return

        start = self._current_page * self._page_size
        end = min(start + self._page_size, total)

        self.table.setColumnCount(len(self._last_columns))
        self.table.setHorizontalHeaderLabels(self._last_columns)
        page_rows = self._last_rows[start:end]
        self.table.setRowCount(len(page_rows))

        for local_row, row_data in enumerate(page_rows):
            global_row = self._global_row(local_row)
            for col_idx, cell_value in enumerate(row_data):
                display_text = _format_display_value(cell_value)
                item = QTableWidgetItem(display_text)
                if cell_value is None:
                    item.setForeground(_NULL_COLOR)
                    item.setFont(self._italic_font)
                else:
                    item.setFont(self._normal_font)

                col_name = self._last_columns[col_idx]
                is_numeric = (
                    _is_numeric_value(cell_value)
                    or self._column_types.get(col_name, "").upper() in _NUMERIC_TYPES_SQL
                )
                if is_numeric:
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

                if (global_row, col_idx) in self._changed_cells:
                    item.setBackground(_CHANGED_COLOR)
                self.table.setItem(local_row, col_idx, item)

        self._resize_columns_smart()
        self._update_nav_buttons()

    def _resize_columns_smart(self):
        header = self.table.horizontalHeader()
        total_width = self.table.viewport().width()
        col_count = self.table.columnCount()
        if col_count == 0:
            return

        self.table.resizeColumnsToContents()

        min_w = 60
        max_w = 400

        for c in range(col_count):
            w = self.table.columnWidth(c)
            if w < min_w:
                self.table.setColumnWidth(c, min_w)
            elif w > max_w:
                self.table.setColumnWidth(c, max_w)

    def _update_nav_buttons(self):
        total = len(self._last_rows)
        self._total_pages = max(1, (total + self._page_size - 1) // self._page_size)
        self.page_label.setText(
            I18N.result_panel["page_label"].format(
                current=self._current_page + 1, total=self._total_pages, count=total
            )
        )
        has_pages = total > self._page_size
        self.page_first_btn.setEnabled(has_pages and self._current_page > 0)
        self.page_prev_btn.setEnabled(has_pages and self._current_page > 0)
        self.page_next_btn.setEnabled(has_pages and self._current_page < self._total_pages - 1)
        self.page_last_btn.setEnabled(has_pages and self._current_page < self._total_pages - 1)

    def show_results(self, columns: list[str], rows: list[list], message: str,
                     editable: bool = False, table_name: str = ""):
        self._last_columns = columns
        self._original_row_count = len(rows)

        if editable and not rows:
            blank = [None] * len(columns)
            rows = [blank]

        self._last_rows = rows
        self._original_rows = copy.deepcopy(rows)
        self._changed_cells.clear()
        self._editable = editable
        self._table_name = table_name
        self._column_types = {}
        if editable:
            self._fetch_column_types()
        self.export_btn.setEnabled(bool(columns))
        self.save_btn.setVisible(editable)

        count = len(rows)
        plural = "s" if count != 1 else ""
        tab_label = I18N.result_panel["tab_editable"] if editable else \
            I18N.result_panel["tab_results_count"].format(count=count, plural=plural)
        self.tabs.setTabText(0, tab_label)

        self.table.setEditTriggers(
            QTableWidget.DoubleClicked if editable else QTableWidget.NoEditTriggers
        )

        self._update_pagination()
        self._go_to_page(0)
        self.message_display.setPlainText(message)
        self.tabs.setCurrentIndex(0)

    def show_message(self, message: str):
        self._clear_edit_state()
        self.export_btn.setEnabled(False)
        self.tabs.setTabText(0, I18N.result_panel["tab_results"])
        self.table.setColumnCount(0)
        self.table.setRowCount(0)
        self.message_display.setPlainText(message)
        self.tabs.setCurrentIndex(1)
        self._update_nav_buttons()

    def show_error(self, error: str):
        self._clear_edit_state()
        self.export_btn.setEnabled(False)
        self.tabs.setTabText(0, I18N.result_panel["tab_results"])
        self.table.setColumnCount(0)
        self.table.setRowCount(0)
        self.message_display.setPlainText(I18N.result_panel["error_prefix"].format(error=error))
        self.message_display.setStyleSheet("""
            QTextEdit {
                background-color: #2d1b1b;
                color: #ff6b6b;
                border: none;
                padding: 8px;
                font-weight: bold;
            }
        """)
        self.tabs.setCurrentIndex(1)
        self._update_nav_buttons()

    def clear(self):
        self._clear_edit_state()
        self.export_btn.setEnabled(False)
        self.tabs.setTabText(0, I18N.result_panel["tab_results"])
        self.table.setColumnCount(0)
        self.table.setRowCount(0)
        self.message_display.clear()
        self.message_display.setStyleSheet("""
            QTextEdit {
                background-color: #1e1e1e;
                color: #d4d4d4;
                border: none;
                padding: 8px;
            }
        """)
        self._update_nav_buttons()

    def _clear_edit_state(self):
        self._last_columns = []
        self._last_rows = []
        self._original_rows = []
        self._changed_cells.clear()
        self._editable = False
        self._table_name = ""
        self._original_row_count = 0
        self._column_types = {}
        self.save_btn.setVisible(False)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)

    def _on_cell_changed(self, local_row: int, col: int):
        if not self._editable:
            return
        global_row = self._global_row(local_row)
        item = self.table.item(local_row, col)
        if item is None:
            return
        current_text = item.text()
        orig_val = self._original_rows[global_row][col]
        orig_text = str(orig_val) if orig_val is not None else "NULL"

        is_changed = (current_text != orig_text)

        if is_changed:
            self._last_rows[global_row][col] = current_text
            self._changed_cells.add((global_row, col))
            item.setBackground(_CHANGED_COLOR)
        else:
            self._changed_cells.discard((global_row, col))
            is_alt = (global_row % 2 == 1) if self.table.alternatingRowColors() else False
            item.setBackground(_ALT_BG if is_alt else _NORMAL_BG)

    def _on_paste(self, start_row: int, start_col: int, data: list[list]):
        if not self._editable:
            return
        total_cols = len(self._last_columns)
        if total_cols == 0:
            return
        pasted = 0

        for r_idx, row_data in enumerate(data):
            global_row = self._global_row(start_row) + r_idx

            while global_row >= len(self._last_rows):
                self._last_rows.append([None] * total_cols)
                self._original_rows.append([None] * total_cols)

            for c_idx, cell_text in enumerate(row_data):
                col = start_col + c_idx
                if col >= total_cols:
                    break
                self._last_rows[global_row][col] = cell_text
                self._changed_cells.add((global_row, col))
                pasted += 1

        if pasted:
            self._update_pagination()
            self._render_page()
            self.tabs.setTabText(0,
                I18N.result_panel["tab_editable"] +
                f" ({I18N.result_panel['paste_done'].format(n=pasted)})"
            )

    def _on_save_changes(self):
        if not self._changed_cells:
            QMessageBox.information(self, I18N.result_panel["save_confirm_title"],
                                    I18N.result_panel["no_changes"])
            return

        rows_affected: dict[int, set[int]] = {}
        for global_row, col in self._changed_cells:
            rows_affected.setdefault(global_row, set()).add(col)

        n_cells = len(self._changed_cells)
        n_rows = len(rows_affected)
        reply = QMessageBox.question(
            self, I18N.result_panel["save_confirm_title"],
            I18N.result_panel["save_confirm_text"].format(n=n_cells, m=n_rows),
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.No:
            return

        if not self._adapter or not self._adapter.is_connected():
            QMessageBox.critical(self, "Erro", "Não conectado ao banco de dados.")
            return

        if not self._table_name:
            QMessageBox.critical(self, "Erro", "Nome da tabela não disponível.")
            return

        conn = None
        try:
            conn = self._adapter.get_connection()
            ok = 0
            errors = 0
            last_error = ""
            inserted_count = 0

            for global_row, cols in rows_affected.items():
                is_insert = global_row >= self._original_row_count

                try:
                    cursor = conn.cursor()

                    if is_insert:
                        col_names = []
                        values = []
                        for col_idx in range(len(self._last_columns)):
                            val = self._last_rows[global_row][col_idx]
                            col_name = self._last_columns[col_idx]
                            if val is not None and str(val).strip() != "":
                                col_names.append(f"[{col_name}]")
                                values.append(self._convert_save_value(col_name, val))
                        if not col_names:
                            cursor.close()
                            continue
                        placeholders = ", ".join("?" for _ in values)
                        sql = (f"INSERT INTO {self._table_name} "
                               f"({', '.join(col_names)}) VALUES ({placeholders})")
                        cursor.execute(sql, values)
                    else:
                        set_parts = []
                        set_values = []
                        where_parts = []
                        where_values = []

                        for col_idx in sorted(cols):
                            col_name = self._last_columns[col_idx]
                            new_val = self._last_rows[global_row][col_idx]
                            set_parts.append(f"[{col_name}] = ?")
                            set_values.append(self._convert_save_value(col_name, new_val))

                        for col_idx in range(len(self._last_columns)):
                            col_name = self._last_columns[col_idx]
                            orig_val = self._original_rows[global_row][col_idx]
                            if orig_val is None:
                                where_parts.append(f"[{col_name}] IS NULL")
                            else:
                                where_parts.append(f"[{col_name}] = ?")
                                where_values.append(self._convert_save_value(col_name, orig_val))

                        sql = (f"UPDATE {self._table_name} SET {', '.join(set_parts)} "
                               f"WHERE {' AND '.join(where_parts)}")
                        cursor.execute(sql, set_values + where_values)

                    conn.commit()
                    cursor.close()
                    ok += 1
                    if is_insert:
                        inserted_count += 1
                except Exception as e:
                    conn.rollback()
                    errors += 1
                    if not last_error:
                        last_error = str(e)[:300]

            if errors == 0:
                for global_row, cols in rows_affected.items():
                    for col in cols:
                        self._original_rows[global_row][col] = self._last_rows[global_row][col]
                self._changed_cells.clear()

                if inserted_count:
                    self._original_row_count = len(self._last_rows)

                self._render_page()
                msg = I18N.result_panel["save_ok"].format(n=ok)
                if inserted_count:
                    msg += f" ({inserted_count} inserida(s))"
                self.message_display.setPlainText(
                    f"> {datetime.now():%H:%M:%S} - {msg}"
                )
                self.status_message.emit(msg, 5000)
                QMessageBox.information(
                    self, I18N.result_panel["save_confirm_title"], msg
                )
            else:
                msg = I18N.result_panel["save_partial"].format(ok=ok, erros=errors)
                if last_error:
                    msg += f"\n\nÚltimo erro: {last_error}"
                self.message_display.setPlainText(
                    f"> {datetime.now():%H:%M:%S} - {msg}"
                )
                self.status_message.emit(msg, 8000)
                QMessageBox.warning(
                    self, I18N.result_panel["save_confirm_title"], msg
                )

        except Exception as e:
            QMessageBox.critical(self, I18N.result_panel["save_confirm_title"],
                                 I18N.result_panel["save_error"].format(msg=str(e)))

    def _export_csv(self):
        if not self._last_columns:
            QMessageBox.information(self, I18N.result_panel["export_title_csv"], I18N.result_panel["no_data_export"])
            return

        default_name = f"query_result_{datetime.now():%Y%m%d_%H%M%S}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self, I18N.result_panel["export_title_csv"], default_name, I18N.result_panel["export_filter_csv"]
        )
        if not file_path:
            return

        try:
            with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(self._last_columns)
                writer.writerows(self._last_rows)
            QMessageBox.information(
                self, I18N.result_panel["export_title_csv"],
                I18N.result_panel["export_ok"].format(n=len(self._last_rows), path=file_path)
            )
            self.status_message.emit(
                I18N.result_panel["export_ok"].format(n=len(self._last_rows), path=file_path), 5000
            )
        except Exception as e:
            QMessageBox.critical(self, I18N.result_panel["export_error"], str(e))

    def _export_excel(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            QMessageBox.critical(self, I18N.result_panel["export_error"],
                                 "Pacote 'openpyxl' necessário para exportar Excel.")
            return

        if not self._last_columns:
            QMessageBox.information(self, I18N.result_panel["export_title_xlsx"], I18N.result_panel["no_data_export"])
            return

        default_name = f"query_result_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, I18N.result_panel["export_title_xlsx"], default_name, I18N.result_panel["export_filter_xlsx"]
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"

            from openpyxl.styles import Font, PatternFill
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0078D4", fill_type="solid")

            for c, col_name in enumerate(self._last_columns, 1):
                cell = ws.cell(row=1, column=c, value=str(col_name))
                cell.font = header_font
                cell.fill = header_fill

            for r, row in enumerate(self._last_rows, 2):
                for c, val in enumerate(row, 1):
                    if val is None:
                        ws.cell(row=r, column=c, value="NULL")
                    elif isinstance(val, (Decimal, float)):
                        ws.cell(row=r, column=c, value=float(val))
                    elif isinstance(val, int):
                        ws.cell(row=r, column=c, value=val)
                    elif isinstance(val, datetime):
                        ws.cell(row=r, column=c, value=val.isoformat() if val else "")
                    else:
                        ws.cell(row=r, column=c, value=str(val))

            wb.save(file_path)
            QMessageBox.information(
                self, I18N.result_panel["export_title_xlsx"],
                I18N.result_panel["export_ok"].format(n=len(self._last_rows), path=file_path)
            )
            self.status_message.emit(
                I18N.result_panel["export_ok"].format(n=len(self._last_rows), path=file_path), 5000
            )
        except Exception as e:
            QMessageBox.critical(self, I18N.result_panel["export_error"], str(e))
